from flask import Flask, request, send_file, jsonify
from flask_cors import CORS
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import PieChart, BarChart, LineChart, Reference
from datetime import datetime, timezone, timedelta
import base64
import hashlib
import hmac
import io
import json
import os
import requests
from urllib.parse import urlparse

app = Flask(__name__)
CORS(app)

# ===== Style helpers =====
def fill(c): return PatternFill('solid', start_color=c)
def border(): return Border(bottom=Side(style='thin', color='2D3250'), right=Side(style='thin', color='2D3250'))
def ca(h='center', v='center', wrap=False): return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
STATUS_COLORS = {'結案':'34D399','轉派技師':'FBBF24','轉派工程師':'A78BFA','客服處理中':'5B8CFF','待派工':'FB923C','待客戶寄回':'F87171'}
BC = ['5B8CFF','7C6CFF','34D399','FBBF24','F87171','FB923C','A78BFA','38BDF8','F472B6']

def set_hdr(ws, row, cols):
    for c, val in enumerate(cols, 1):
        cell = ws.cell(row=row, column=c, value=val)
        cell.font = Font(name='Arial', bold=True, color='94A3B8', size=10)
        cell.fill = fill('2D3250')
        cell.alignment = ca()
        cell.border = border()

def title_row(ws, row, text, ncols, bg='5B8CFF'):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name='Arial', bold=True, color='FFFFFF', size=12)
    c.fill = fill(bg)
    c.alignment = ca('left')
    ws.row_dimensions[row].height = 22

def calc_dur(s, e):
    if not s or not e: return ''
    try:
        ms = (datetime.fromisoformat(e) - datetime.fromisoformat(s)).total_seconds()
        if ms <= 0: return ''
        h = int(ms // 3600)
        m = int((ms % 3600) // 60)
        if h >= 24:
            d = h // 24; rh = h % 24
            return f'{d}天{rh}小時' if rh else f'{d}天'
        return f'{h}小時{m}分' if h else f'{m}分鐘'
    except: return ''

def fmt_dt(value):
    if not value:
        return ''
    return str(value).replace('T', ' ')[:16]

def build_status_focus(rows):
    grouped = {}
    for r in rows:
        company = r.get('company') or '未填公司'
        grouped.setdefault(company, []).append(r)

    lines = []
    status = rows[0].get('status') if rows else ''
    for company in sorted(grouped.keys()):
        company_rows = sorted(grouped[company], key=lambda item: item.get('date') or '')
        problem_counts = {}
        for r in company_rows:
            subcategory = r.get('subcategory') or '未填問題'
            problem_counts[subcategory] = problem_counts.get(subcategory, 0) + 1

        problems = []
        for problem, count in problem_counts.items():
            problems.append(f'{problem}（{count}筆）' if count > 1 else problem)

        company_label = company
        if status != '結案':
            company_label = f'{company}（{len(company_rows)}筆）'
        lines.append(f'{company_label}：' + '、'.join(problems))
    return '\n'.join(lines)

def build_incoming_time_summary(rows):
    dates = sorted(fmt_dt(r.get('date')) for r in rows if r.get('date'))
    if not dates:
        return ''
    if len(dates) == 1:
        return f'進線：{dates[0]}\n共 1 筆'
    return f'最早：{dates[0]}\n最新：{dates[-1]}\n共 {len(rows)} 筆'

def is_dispatch_overdue(r):
    if r.get('status') == '結案': return False
    if not r.get('date'): return False
    if r.get('handler') == '客戶': return False
    try: return (datetime.now() - datetime.fromisoformat(r['date'])).days > 7
    except: return False

def parse_report_date(value):
    if not value:
        return None
    try:
        return datetime.fromisoformat(str(value)[:10]).date()
    except:
        return None

def get_record_date(value):
    if not value:
        return None
    try:
        return datetime.fromisoformat(str(value)).date()
    except:
        try:
            return datetime.fromisoformat(str(value)[:10]).date()
        except:
            return None

def build_previous_period_records(all_records, from_date, to_date):
    start = parse_report_date(from_date)
    end = parse_report_date(to_date)
    if not start or not end or end < start:
        return []
    days = (end - start).days + 1
    prev_end = start - timedelta(days=1)
    prev_start = prev_end - timedelta(days=days - 1)
    rows = []
    for r in all_records:
        record_date = get_record_date(r.get('date'))
        if record_date and prev_start <= record_date <= prev_end:
            rows.append(r)
    return rows

def build_category_counts(records):
    counts = {}
    for r in records:
        category = r.get('category') or '其他'
        counts[category] = counts.get(category, 0) + 1
    return counts

def format_weekly_category_delta(current_count, previous_count):
    diff = current_count - previous_count
    if diff > 0:
        return f'↑{diff}件（上週 {previous_count}件）'
    if diff < 0:
        return f'↓{abs(diff)}件（上週 {previous_count}件）'
    return f'持平（上週 {previous_count}件）'

def get_report_end_datetime(to_date):
    if not to_date:
        return datetime.now()
    try:
        text = str(to_date)
        if len(text) == 10:
            return datetime.fromisoformat(text + 'T23:59:59')
        return datetime.fromisoformat(text)
    except:
        return datetime.now()

def get_elapsed_days_as_of(r, end_dt):
    if not r.get('date'):
        return 0
    try:
        return max((end_dt - datetime.fromisoformat(r['date'])).days, 0)
    except:
        return 0

def is_overdue_as_of(r, end_dt):
    if r.get('status') == '結案': return False
    if not r.get('date'): return False
    if r.get('handler') == '客戶': return False
    return get_elapsed_days_as_of(r, end_dt) > 7

def map_product(p):
    if not p: return '其他'
    if 'DMVR' in p: return 'FMS-DMVR'
    if 'GPS' in p: return 'FMS-GPS'
    if '冷鏈' in p: return 'FMS-冷鏈'
    if '雷達' in p: return 'FMS-雷達'
    return '其他'

def is_parent(r):
    """判斷是否為父單（編號格式 YYYYMMDD-NNN，沒有第三段）"""
    parts = r.get('id', '').split('-')
    return len(parts) == 2

# ===== 超過7天未結案追蹤區段（週報 ③ 下方用）=====
def get_wait_reference_value(r):
    return r.get('date') or ''

def get_wait_days(r):
    ref = get_wait_reference_value(r)
    if not ref:
        return 0
    try:
        return (datetime.now() - datetime.fromisoformat(ref)).days
    except:
        return 0

def write_all_open_section(ws, start_row, all_records):
    """在指定 row 寫入截至今日超過 7 天未結案追蹤，回傳結束後的 row。"""
    overdue_cases = [r for r in all_records if is_dispatch_overdue(r)]

    ws.row_dimensions[start_row].height = 10
    row = start_row + 1

    if not overdue_cases:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        c = ws.cell(row=row, column=1, value='✅ 目前無超過7天未結案案件')
        c.font = Font(name='Arial', bold=True, color='34D399', size=11)
        c.alignment = ca()
        return row + 1

    title_row(ws, row, f'📌 超過7天未結案追蹤（截至今日，全期間共 {len(overdue_cases)} 筆）', 8, bg='2D3250')
    row += 1

    set_hdr(ws, row, ['已等待', '進線日期', '公司名稱', '車牌', '問題次分類', '處理狀態', '負責人員', '備註說明'])
    for col, w in [('A',10),('B',14),('C',16),('D',16),('E',26),('F',14),('G',12),('H',32)]:
        ws.column_dimensions[col].width = w
    row += 1

    for r in sorted(overdue_cases, key=lambda x: get_wait_days(x), reverse=True):
        _write_open_row(ws, row, r, is_child=not is_parent(r))
        row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    c = ws.cell(row=row, column=1, value='🔴  逾7天（高風險）')
    c.font = Font(name='Arial', bold=True, color='F87171', size=10)
    c.fill = fill('1E2235')
    c.alignment = ca('left')
    ws.row_dimensions[row].height = 18
    return row + 1

def _write_open_row(ws, row, r, is_child=False):
    """寫入一筆未結案記錄"""
    try:
        wait_days = (datetime.now() - datetime.fromisoformat(r['date'])).days if r.get('date') else 0
    except:
        wait_days = 0

    wait_str = f'{wait_days}天'
    wait_color = 'F87171' if wait_days > 14 else ('FB923C' if wait_days > 7 else 'E2E8F0')
    bg = '1A1D27' if is_child else '161925'
    id_prefix = '  ↳ ' if is_child else ''

    vals = [
        id_prefix + r.get('id', ''),
        r.get('company', ''),
        r.get('plate', ''),
        r.get('subcategory', ''),
        r.get('handler', '—'),
        r.get('date', '')[:10] if r.get('date') else '',
        wait_str
    ]
    colors = ['A78BFA' if is_child else 'E2E8F0', 'FFFFFF', '94A3B8', '94A3B8', 'E2E8F0', '94A3B8', wait_color]
    bolds = [True, False, False, False, False, False, True]

    for c2, (val, color, bold) in enumerate(zip(vals, colors, bolds), 1):
        c = ws.cell(row=row, column=c2, value=val)
        c.font = Font(name='Arial', bold=bold, color=color, size=10)
        c.fill = fill(bg)
        c.alignment = ca('left') if c2 <= 4 else ca()
        c.border = border()
    ws.row_dimensions[row].height = 16


def _write_open_row(ws, row, r, is_child=False):
    wait_days = get_wait_days(r)
    wait_str = f'{wait_days}天'
    wait_color = 'F87171' if wait_days >= 7 else ('FB923C' if wait_days >= 3 else 'E2E8F0')
    bg = '1A1D27' if is_child else '161925'

    wait_reference = get_wait_reference_value(r)
    wait_date = fmt_dt(wait_reference)[:10]

    vals = [
        wait_str,
        wait_date,
        r.get('company', ''),
        r.get('plate', ''),
        r.get('subcategory', ''),
        r.get('status', ''),
        r.get('handler', '—'),
        r.get('result') or r.get('description') or ''
    ]
    colors = [wait_color, '94A3B8', 'FFFFFF', '94A3B8', '94A3B8', STATUS_COLORS.get(r.get('status',''),'E2E8F0'), 'E2E8F0', '94A3B8']
    bolds = [True, False, False, False, False, True, False, False]

    for c2, (val, color, bold) in enumerate(zip(vals, colors, bolds), 1):
        c = ws.cell(row=row, column=c2, value=val)
        c.font = Font(name='Arial', bold=bold, color=color, size=10)
        c.fill = fill(bg)
        c.alignment = ca('left', wrap=True) if c2 in (3, 5, 8) else ca()
        c.border = border()
    ws.row_dimensions[row].height = 20

# ===== WEEKLY REPORT =====
def generate_weekly(records, from_date, to_date, all_records=None):
    if all_records is None:
        all_records = records

    total = len(records)
    closed = sum(1 for r in records if r.get('status') == '結案')
    open_cnt = total - closed
    close_rate = f'{closed/total*100:.1f}%' if total else '0%'
    od = [r for r in all_records if is_dispatch_overdue(r)]
    previous_records = build_previous_period_records(all_records, from_date, to_date)
    previous_category_counts = build_category_counts(previous_records)
    durs = []
    for r in records:
        if r.get('status') == '結案' and r.get('date') and r.get('closeDate'):
            try:
                sec = (datetime.fromisoformat(r['closeDate']) - datetime.fromisoformat(r['date'])).total_seconds()
                if sec > 0: durs.append(sec)
            except: pass
    avg_h = round(sum(durs)/len(durs)/3600) if durs else 0
    avg_str = f'{avg_h//24}天{avg_h%24}小時' if avg_h >= 24 else f'{avg_h}小時'
    label = f'{from_date} ～ {to_date}'
    wb = Workbook()

    # ===== 封面 =====
    ws0 = wb.active
    ws0.title = '封面摘要'
    ws0.sheet_view.showGridLines = False
    for col, w in [('A',3),('B',22),('C',18),('D',18),('E',18),('F',18),('G',3)]:
        ws0.column_dimensions[col].width = w

    ws0.merge_cells('B2:F2')
    c = ws0.cell(row=2, column=2, value='⚡ 售服案件週報')
    c.font = Font(name='Arial', bold=True, color='FFFFFF', size=18)
    c.fill = fill('1A1D27'); c.alignment = ca('left')
    ws0.row_dimensions[2].height = 36

    ws0.merge_cells('B3:F3')
    c = ws0.cell(row=3, column=2, value=f'報告期間：{label}')
    c.font = Font(name='Arial', color='94A3B8', size=11)
    c.fill = fill('1A1D27'); c.alignment = ca('left')
    ws0.row_dimensions[3].height = 22
    ws0.row_dimensions[4].height = 12

    kpis = [('進線總件數',total,'5B8CFF','件'),('已結案',closed,'34D399',f'結案率 {close_rate}'),('未結案',open_cnt,'FBBF24','件'),('逾7天未結案',len(od),'F87171','件')]
    for i, (lbl, val, color, sub) in enumerate(kpis):
        col = 2 + i
        c = ws0.cell(row=5, column=col, value=lbl)
        c.font = Font(name='Arial', color='94A3B8', size=9); c.fill = fill('22263A'); c.alignment = ca()
        ws0.row_dimensions[5].height = 14
        ws0.merge_cells(start_row=6, start_column=col, end_row=7, end_column=col)
        c = ws0.cell(row=6, column=col, value=val)
        c.font = Font(name='Arial', bold=True, color=color, size=26); c.fill = fill('22263A'); c.alignment = ca()
        ws0.row_dimensions[6].height = 28; ws0.row_dimensions[7].height = 10
        c = ws0.cell(row=8, column=col, value=sub)
        c.font = Font(name='Arial', color='64748B', size=9); c.fill = fill('22263A'); c.alignment = ca()
        ws0.row_dimensions[8].height = 14

    ws0.row_dimensions[9].height = 12
    ws0.merge_cells('B10:F10')
    c = ws0.cell(row=10, column=2, value='👤 本週人員負責件數')
    c.font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    c.fill = fill('2D3250'); c.alignment = ca('left')
    ws0.row_dimensions[10].height = 18

    handler_c = {}
    for r in records:
        if r.get('handler'): handler_c[r['handler']] = handler_c.get(r['handler'], 0) + 1
    row = 11
    for h, cnt in sorted(handler_c.items(), key=lambda x: -x[1]):
        bg = '1E2235' if row%2==0 else '161925'
        for col, val in [(2, h),(3, f'{cnt} 件')]:
            c = ws0.cell(row=row, column=col, value=val)
            c.font = Font(name='Arial', size=10, color='E2E8F0')
            c.fill = fill(bg); c.alignment = ca('left') if col==2 else ca()
            c.border = border()
        ws0.row_dimensions[row].height = 16; row += 1

    # ===== ① 進線管道 =====
    ws1 = wb.create_sheet('① 進線管道分析')
    ws1.sheet_view.showGridLines = False
    title_row(ws1, 1, f'📡 進線管道分析　｜　{label}', 3)
    set_hdr(ws1, 2, ['進線管道','件數','佔比'])
    ws1.column_dimensions['A'].width = 18; ws1.column_dimensions['B'].width = 10; ws1.column_dimensions['C'].width = 12

    ch_c = {}
    for r in records: ch_c[r.get('channel') or '未知'] = ch_c.get(r.get('channel') or '未知', 0) + 1
    row = 3
    for ch, cnt in sorted(ch_c.items(), key=lambda x: -x[1]):
        bg = '1E2235' if row%2==0 else '161925'
        for c2, (val, color, bold) in enumerate([(ch,'E2E8F0',False),(cnt,'5B8CFF',True),(f'{cnt/total*100:.1f}%' if total else '0%','94A3B8',False)], 1):
            c = ws1.cell(row=row, column=c2, value=val)
            c.font = Font(name='Arial', bold=bold, color=color, size=10)
            c.fill = fill(bg); c.alignment = ca(); c.border = border()
        row += 1
    for c2, val in enumerate(['合計',total,'100%'], 1):
        c = ws1.cell(row=row, column=c2, value=val)
        c.font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
        c.fill = fill('2D3250'); c.alignment = ca(); c.border = border()
    pie = PieChart(); pie.title='進線管道佔比'; pie.style=10; pie.width=14; pie.height=10
    labels = Reference(ws1, min_col=1, min_row=3, max_row=2+len(ch_c))
    data = Reference(ws1, min_col=2, min_row=2, max_row=2+len(ch_c))
    pie.add_data(data, titles_from_data=True); pie.set_categories(labels)
    ws1.add_chart(pie, 'E2')

    # ===== ② 問題分類 =====
    ws2 = wb.create_sheet('② 問題分類統計')
    ws2.sheet_view.showGridLines = False
    title_row(ws2, 1, f'🏷️ 問題分類統計　｜　{label}', 4)
    set_hdr(ws2, 2, ['問題大類','總件數','主要次分類','與上週差異'])
    ws2.column_dimensions['A'].width = 14; ws2.column_dimensions['B'].width = 10
    ws2.column_dimensions['C'].width = 50; ws2.column_dimensions['D'].width = 22

    cat_map = {}
    for r in records:
        k = r.get('category') or '其他'
        if k not in cat_map: cat_map[k] = {'total':0,'subs':{}}
        cat_map[k]['total'] += 1
        s = r.get('subcategory') or '其他'
        cat_map[k]['subs'][s] = cat_map[k]['subs'].get(s,0)+1
    sorted_cats = sorted(cat_map.items(), key=lambda x: -x[1]['total'])
    row = 3
    for cat, v in sorted_cats:
        top = '、'.join(f"{s}({n})" for s,n in sorted(v['subs'].items(), key=lambda x:-x[1])[:4])
        delta = format_weekly_category_delta(v['total'], previous_category_counts.get(cat, 0))
        bg = '1E2235' if row%2==0 else '161925'
        for c2,(val,color,bold) in enumerate([(cat,'E2E8F0',False),(v['total'],'5B8CFF',True),(top,'94A3B8',False),(delta,'FBBF24',False)], 1):
            c = ws2.cell(row=row, column=c2, value=val)
            c.font = Font(name='Arial', bold=bold, color=color, size=10)
            c.fill = fill(bg); c.alignment = ca() if c2<=2 else ca('left',wrap=True); c.border = border()
        ws2.row_dimensions[row].height = 16; row += 1
    bar = BarChart(); bar.type='bar'; bar.title='問題大類件數'; bar.style=10; bar.width=16; bar.height=12
    cats_r = Reference(ws2, min_col=1, min_row=3, max_row=2+len(sorted_cats))
    data_r = Reference(ws2, min_col=2, min_row=2, max_row=2+len(sorted_cats))
    bar.add_data(data_r, titles_from_data=True); bar.set_categories(cats_r)
    ws2.add_chart(bar, 'E2')

    # ===== ③ 處理狀態 + 超過7天未結案追蹤 =====
    ws3 = wb.create_sheet('③ 處理狀態總覽')
    ws3.sheet_view.showGridLines = False
    title_row(ws3, 1, f'📋 處理狀態總覽　｜　{label}', 8)
    set_hdr(ws3, 2, ['處理狀態','件數','處理人員','重點說明'])
    ws3.column_dimensions['A'].width = 14; ws3.column_dimensions['B'].width = 8
    ws3.column_dimensions['C'].width = 24; ws3.column_dimensions['D'].width = 72

    status_groups = {}
    for r in records:
        k = r.get('status') or '未知'
        status_groups.setdefault(k, []).append(r)
    row = 3
    for st, rows in sorted(status_groups.items(), key=lambda x: -len(x[1])):
        handlers = '、'.join(set(r['handler'] for r in rows if r.get('handler')))
        notes = build_status_focus(rows)
        bg = '1E2235' if row%2==0 else '161925'
        for c2, val in enumerate([st, len(rows), handlers, notes], 1):
            c = ws3.cell(row=row, column=c2, value=val)
            c.font = Font(name='Arial', bold=(c2==1), color=STATUS_COLORS.get(st,'E2E8F0') if c2==1 else 'E2E8F0', size=10)
            c.fill = fill(bg); c.alignment = ca() if c2<=2 else ca('left',wrap=True); c.border = border()
        ws3.row_dimensions[row].height = max(45, len(notes.split('\n'))*22); row += 1

    # 圓餅圖
    chart_row = row + 1
    ws3.cell(row=chart_row, column=1, value='狀態'); ws3.cell(row=chart_row, column=2, value='件數')
    for i,(st,rows) in enumerate(sorted(status_groups.items(), key=lambda x:-len(x[1])),1):
        ws3.cell(row=chart_row+i, column=1, value=st); ws3.cell(row=chart_row+i, column=2, value=len(rows))
    pie2 = PieChart(); pie2.title='處理狀態分佈'; pie2.style=10; pie2.width=14; pie2.height=10
    lb2 = Reference(ws3, min_col=1, min_row=chart_row+1, max_row=chart_row+len(status_groups))
    d2 = Reference(ws3, min_col=2, min_row=chart_row, max_row=chart_row+len(status_groups))
    pie2.add_data(d2, titles_from_data=True); pie2.set_categories(lb2)
    ws3.add_chart(pie2, 'F2')

    # ===== 超過7天未結案追蹤（接在圓餅圖資料後） =====
    section_start = chart_row + len(status_groups) + 3
    write_all_open_section(ws3, section_start, all_records)

    # ===== 說明 =====
    ws4 = wb.create_sheet('📖 說明')
    ws4.sheet_view.showGridLines = False
    ws4.column_dimensions['A'].width = 24
    ws4.column_dimensions['B'].width = 72
    title_row(ws4, 1, '📖 週報說明文件', 2, bg='2D3250')

    explanation_rows = [
        (3, '📌 案件編號格式', ''),
        (4, '一般案件', 'YYYYMMDD-流水號 / 例：20260522-001'),
        (5, '子案件', '母案件編號-子流水號 / 例：20260422-003-02'),
        (7, '🔄 處理狀態定義', ''),
        (8, '客服處理中', '客服人員正在跟進，尚未派工'),
        (9, '待派工', '已確認需派技師或工程師，尚未指派人員'),
        (10, '待客戶確認', '已提供報價或方案，等待客戶回覆確認'),
        (11, '轉派技師', '已指派外部技師前往現場處理'),
        (12, '轉派工程師', '已指派內部工程師深入處理'),
        (13, '待客戶寄回', '需客戶將設備寄回公司'),
        (14, '結案', '問題已解決並確認，案件關閉'),
        (16, '⚠️ 超過7天未結案計算邏輯', ''),
        (17, '計算基準', '從「進線日期時間」起算'),
        (18, '篩選條件', '狀態不是「結案」，且已等待 > 7 天'),
        (19, '範圍', '全期間所有案件，不限本週進線'),
        (20, '排除條件', '負責人為「客戶」的案件不列入'),
        (22, '🎨 顏色說明（Sheet③ 追蹤表）', ''),
        (23, '🔴 紅底', '已等待 >= 7 天，高風險'),
        (24, '🟡 黃底', '已等待 3~6 天，追蹤中'),
        (25, '⬜ 白/灰', '已等待 0~2 天，正常'),
    ]

    for row, left, right in explanation_rows:
        if right == '':
            ws4.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
            c = ws4.cell(row=row, column=1, value=left)
            c.font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
            c.fill = fill('2D3250')
            c.alignment = ca('left')
        else:
            for col, val in [(1, left), (2, right)]:
                c = ws4.cell(row=row, column=col, value=val)
                c.font = Font(name='Arial', bold=(col == 1), color='E2E8F0', size=10)
                c.fill = fill('161925' if row % 2 else '1E2235')
                c.alignment = ca('left', wrap=True)
                c.border = border()
        ws4.row_dimensions[row].height = 20

    return wb

# ===== MONTHLY REPORT =====
def generate_monthly(records, from_date, to_date):
    total = len(records)
    closed = sum(1 for r in records if r.get('status') == '結案')
    open_cnt = total - closed
    close_rate = f'{closed/total*100:.1f}%' if total else '0%'
    report_end = get_report_end_datetime(to_date)
    od = [r for r in records if is_overdue_as_of(r, report_end)]
    durs = []
    for r in records:
        if r.get('status') == '結案' and r.get('date') and r.get('closeDate'):
            try:
                sec = (datetime.fromisoformat(r['closeDate'])-datetime.fromisoformat(r['date'])).total_seconds()
                if sec > 0: durs.append(sec)
            except: pass
    avg_h = round(sum(durs)/len(durs)/3600) if durs else 0
    avg_str = f'{avg_h//24}天{avg_h%24}小時' if avg_h >= 24 else f'{avg_h}小時'
    label = f'{from_date} ～ {to_date}'
    wb = Workbook()

    # ===== 封面 =====
    ws0 = wb.active; ws0.title = '封面摘要'
    ws0.sheet_view.showGridLines = False
    for col, w in [('A',3),('B',22),('C',18),('D',18),('E',18),('F',18),('G',3)]:
        ws0.column_dimensions[col].width = w
    ws0.merge_cells('B2:F2')
    c = ws0.cell(row=2, column=2, value='⚡ 售服案件月報')
    c.font = Font(name='Arial', bold=True, color='FFFFFF', size=18)
    c.fill = fill('1A1D27'); c.alignment = ca('left'); ws0.row_dimensions[2].height = 36
    ws0.merge_cells('B3:F3')
    c = ws0.cell(row=3, column=2, value=f'報告期間：{label}　｜　產製日期：{to_date}')
    c.font = Font(name='Arial', color='94A3B8', size=11)
    c.fill = fill('1A1D27'); c.alignment = ca('left'); ws0.row_dimensions[3].height = 22
    ws0.row_dimensions[4].height = 12

    kpis = [('進線總件數',total,'5B8CFF','件'),('已結案',closed,'34D399',f'結案率 {close_rate}'),('未結案',open_cnt,'FBBF24','件'),('逾7天未結案',len(od),'F87171','件')]
    for i,(lbl,val,color,sub) in enumerate(kpis):
        col = 2+i
        c = ws0.cell(row=5, column=col, value=lbl)
        c.font = Font(name='Arial', color='94A3B8', size=9); c.fill = fill('22263A'); c.alignment = ca()
        ws0.row_dimensions[5].height = 14
        ws0.merge_cells(start_row=6, start_column=col, end_row=7, end_column=col)
        c = ws0.cell(row=6, column=col, value=val)
        c.font = Font(name='Arial', bold=True, color=color, size=22); c.fill = fill('22263A'); c.alignment = ca()
        ws0.row_dimensions[6].height = 24; ws0.row_dimensions[7].height = 10
        c = ws0.cell(row=8, column=col, value=sub)
        c.font = Font(name='Arial', color='64748B', size=9); c.fill = fill('22263A'); c.alignment = ca()
        ws0.row_dimensions[8].height = 14
    ws0.row_dimensions[9].height = 12
    kpis2 = [('結案率',close_rate,'A78BFA'),('平均處理時間',avg_str,'38BDF8'),('最長逾期天數',f'{max((get_elapsed_days_as_of(r, report_end) for r in od if r.get("date")), default=0)}天','FB923C')]
    for i,(lbl,val,color) in enumerate(kpis2):
        col = 2+i
        c = ws0.cell(row=10, column=col, value=lbl)
        c.font = Font(name='Arial', color='94A3B8', size=9); c.fill = fill('1A1D27'); c.alignment = ca()
        ws0.row_dimensions[10].height = 14
        ws0.merge_cells(start_row=11, start_column=col, end_row=12, end_column=col)
        c = ws0.cell(row=11, column=col, value=val)
        c.font = Font(name='Arial', bold=True, color=color, size=18); c.fill = fill('1A1D27'); c.alignment = ca()
        ws0.row_dimensions[11].height = 24; ws0.row_dimensions[12].height = 14
    ws0.row_dimensions[13].height = 16
    ws0.merge_cells('B14:F14')
    c = ws0.cell(row=14, column=2, value='👤 本月人員負責件數')
    c.font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    c.fill = fill('2D3250'); c.alignment = ca('left'); ws0.row_dimensions[14].height = 18
    handler_c = {}
    for r in records:
        if r.get('handler'): handler_c[r['handler']] = handler_c.get(r['handler'], 0)+1
    row = 15
    for h, cnt in sorted(handler_c.items(), key=lambda x:-x[1]):
        bg = '1E2235' if row%2==0 else '161925'
        for col, val in [(2,h),(3,f'{cnt} 件')]:
            c = ws0.cell(row=row, column=col, value=val)
            c.font = Font(name='Arial', size=10, color='E2E8F0')
            c.fill = fill(bg); c.alignment = ca('left') if col==2 else ca(); c.border = border()
        ws0.row_dimensions[row].height = 16; row += 1

    # ===== ① 進線管道 =====
    ws1 = wb.create_sheet('① 進線管道分析')
    ws1.sheet_view.showGridLines = False
    title_row(ws1, 1, f'📡 進線管道分析　｜　{label}', 3)
    set_hdr(ws1, 2, ['進線管道','件數','佔比'])
    ws1.column_dimensions['A'].width = 18; ws1.column_dimensions['B'].width = 12; ws1.column_dimensions['C'].width = 12
    ch_c = {}
    for r in records: ch_c[r.get('channel') or '未知'] = ch_c.get(r.get('channel') or '未知',0)+1
    row = 3
    for ch, cnt in sorted(ch_c.items(), key=lambda x:-x[1]):
        bg = '1E2235' if row%2==0 else '161925'
        for c2,(val,color,bold) in enumerate([(ch,'E2E8F0',False),(cnt,'5B8CFF',True),(f'{cnt/total*100:.1f}%' if total else '0%','94A3B8',False)],1):
            c = ws1.cell(row=row, column=c2, value=val)
            c.font = Font(name='Arial', bold=bold, color=color, size=10)
            c.fill = fill(bg); c.alignment = ca(); c.border = border()
        row += 1
    for c2, val in enumerate(['合計',total,'100%'],1):
        c = ws1.cell(row=row, column=c2, value=val)
        c.font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
        c.fill = fill('2D3250'); c.alignment = ca(); c.border = border()
    pie = PieChart(); pie.title='進線管道佔比'; pie.style=10; pie.width=14; pie.height=10
    lb = Reference(ws1, min_col=1, min_row=3, max_row=2+len(ch_c))
    dt = Reference(ws1, min_col=2, min_row=2, max_row=2+len(ch_c))
    pie.add_data(dt, titles_from_data=True); pie.set_categories(lb)
    ws1.add_chart(pie, 'E2')

    # ===== ② 案件類別 =====
    ws2 = wb.create_sheet('② 案件類別分析')
    ws2.sheet_view.showGridLines = False
    title_row(ws2, 1, f'📦 案件類別分析（產品別）　｜　{label}', 3)
    set_hdr(ws2, 2, ['產品類別','件數','佔比'])
    ws2.column_dimensions['A'].width = 16; ws2.column_dimensions['B'].width = 10; ws2.column_dimensions['C'].width = 12
    prod_map = {'FMS-GPS':0,'FMS-DMVR':0,'FMS-冷鏈':0,'FMS-雷達':0,'其他':0}
    for r in records: prod_map[map_product(r.get('product',''))] += 1
    prod_list = [(k,v) for k,v in prod_map.items() if v>0]
    prod_list.sort(key=lambda x:-x[1])
    total_prod = sum(v for _,v in prod_list)
    PROD_COLORS = ['5B8CFF','34D399','FBBF24','A78BFA','94A3B8']
    row = 3
    for i,(prod,cnt) in enumerate(prod_list):
        bg = '1E2235' if row%2==0 else '161925'
        for c2,(val,color,bold) in enumerate([(prod,PROD_COLORS[i%5],True),(cnt,'FFFFFF',True),(f'{cnt/total_prod*100:.1f}%' if total_prod else '0%','94A3B8',False)],1):
            c = ws2.cell(row=row, column=c2, value=val)
            c.font = Font(name='Arial', bold=bold, color=color, size=10)
            c.fill = fill(bg); c.alignment = ca(); c.border = border()
        ws2.row_dimensions[row].height = 18; row += 1
    for c2,(val,color) in enumerate([('合計','FFFFFF'),(total_prod,'FFFFFF'),('100%','FFFFFF')],1):
        c = ws2.cell(row=row, column=c2, value=val)
        c.font = Font(name='Arial', bold=True, color=color, size=10)
        c.fill = fill('2D3250'); c.alignment = ca(); c.border = border()
    chart_row = row+2
    ws2.cell(row=chart_row, column=1, value='產品類別'); ws2.cell(row=chart_row, column=2, value='件數')
    for i2,(prod,cnt) in enumerate(prod_list,1):
        ws2.cell(row=chart_row+i2, column=1, value=prod); ws2.cell(row=chart_row+i2, column=2, value=cnt)
    pie2 = PieChart(); pie2.title='產品類別佔比'; pie2.style=10; pie2.width=16; pie2.height=12
    lb2 = Reference(ws2, min_col=1, min_row=chart_row+1, max_row=chart_row+len(prod_list))
    d2 = Reference(ws2, min_col=2, min_row=chart_row, max_row=chart_row+len(prod_list))
    pie2.add_data(d2, titles_from_data=True); pie2.set_categories(lb2)
    ws2.add_chart(pie2, 'E2')

    # ===== ③ 客戶進線 TOP5 =====
    ws3 = wb.create_sheet('③ 客戶進線排行')
    ws3.sheet_view.showGridLines = False
    title_row(ws3, 1, f'🏢 客戶進線排行 TOP5　｜　{label}', 4)
    set_hdr(ws3, 2, ['排名','客戶名稱','件數','主要問題'])
    ws3.column_dimensions['A'].width = 8; ws3.column_dimensions['B'].width = 16
    ws3.column_dimensions['C'].width = 10; ws3.column_dimensions['D'].width = 45
    company_c = {}
    company_issues = {}
    for r in records:
        co = r.get('company','')
        if co:
            company_c[co] = company_c.get(co,0)+1
            company_issues.setdefault(co, {})
            cat = r.get('category','其他')
            company_issues[co][cat] = company_issues[co].get(cat,0)+1
    top5 = sorted(company_c.items(), key=lambda x:-x[1])[:5]
    rank_colors = ['FFD700','C0C0C0','CD7F32','E2E8F0','E2E8F0']
    row = 3
    for rank,(co,cnt) in enumerate(top5,1):
        issues = '、'.join(f"{k}({v})" for k,v in sorted(company_issues.get(co,{}).items(), key=lambda x:-x[1])[:2])
        bg = '1E2235' if row%2==0 else '161925'
        for c2,(val,color,bold) in enumerate([(rank,rank_colors[rank-1],rank<=3),(co,'FFFFFF',False),(cnt,'5B8CFF',True),(issues,'94A3B8',False)],1):
            c = ws3.cell(row=row, column=c2, value=val)
            c.font = Font(name='Arial', bold=bold, color=color, size=10)
            c.fill = fill(bg); c.alignment = ca() if c2!=4 else ca('left',wrap=True); c.border = border()
        ws3.row_dimensions[row].height = 16; row += 1
    bar2 = BarChart(); bar2.type='bar'; bar2.title='客戶進線件數 TOP5'; bar2.style=10; bar2.width=16; bar2.height=10
    cats_r = Reference(ws3, min_col=2, min_row=3, max_row=2+len(top5))
    data_r = Reference(ws3, min_col=3, min_row=2, max_row=2+len(top5))
    bar2.add_data(data_r, titles_from_data=True); bar2.set_categories(cats_r)
    ws3.add_chart(bar2, 'F2')

    # ===== ④ 逾7天未結案 =====
    ws4 = wb.create_sheet('④ 逾7天未結案')
    ws4.sheet_view.showGridLines = False
    max_days = max([get_elapsed_days_as_of(r, report_end) for r in od if r.get('date')], default=0)
    ws4.merge_cells('A1:G1')
    sc = ws4.cell(row=1, column=1, value=f'⚠️  截至月底共 {len(od)} 筆超過7天未結案　｜　最長已逾 {max_days} 天　｜　{label}')
    sc.font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    sc.fill = fill('7F1D1D'); sc.alignment = ca('left'); ws4.row_dimensions[1].height = 22
    set_hdr(ws4, 2, ['進線日期時間','車牌','公司名稱','問題次分類','處理狀態','負責人員','已逾天數'])
    for col, w in [('A',16),('B',16),('C',14),('D',26),('E',14),('F',12),('G',14)]:
        ws4.column_dimensions[col].width = w
    row = 3
    for r in sorted(od, key=lambda x: x.get('date','')):
        days = get_elapsed_days_as_of(r, report_end)
        day_color = 'F87171' if days>14 else 'FB923C'
        bg = '2A1515' if row%2==0 else '221212'
        vals = [fmt_dt(r.get('date')),r.get('plate',''),r.get('company',''),r.get('subcategory',''),r.get('status',''),r.get('handler','—'),f'{days}天']
        colors = ['94A3B8','94A3B8','FFFFFF','94A3B8',STATUS_COLORS.get(r.get('status',''),'E2E8F0'),'E2E8F0',day_color]
        for c2,(val,color) in enumerate(zip(vals,colors),1):
            c = ws4.cell(row=row, column=c2, value=val)
            c.font = Font(name='Arial', bold=(c2==7), color=color, size=10)
            c.fill = fill(bg); c.alignment = ca(); c.border = border()
        ws4.row_dimensions[row].height = 16; row += 1
    if not od:
        ws4.merge_cells('A3:G3')
        c = ws4.cell(row=3, column=1, value='✅ 本月無逾期未結案')
        c.font = Font(name='Arial', bold=True, color='34D399', size=12); c.alignment = ca()

    # ===== ⑤ 問題分類 =====
    ws5 = wb.create_sheet('⑤ 問題分類統計')
    ws5.sheet_view.showGridLines = False
    title_row(ws5, 1, f'🏷️ 問題分類統計　｜　{label}', 3)
    set_hdr(ws5, 2, ['問題大類','總件數','主要次分類（件數）'])
    ws5.column_dimensions['A'].width = 14; ws5.column_dimensions['B'].width = 10; ws5.column_dimensions['C'].width = 50
    cat_map = {}
    for r in records:
        k = r.get('category') or '其他'
        if k not in cat_map: cat_map[k] = {'total':0,'subs':{}}
        cat_map[k]['total'] += 1
        s = r.get('subcategory') or '其他'
        cat_map[k]['subs'][s] = cat_map[k]['subs'].get(s,0)+1
    sorted_cats = sorted(cat_map.items(), key=lambda x:-x[1]['total'])
    row = 3
    for cat, v in sorted_cats:
        top = '、'.join(f"{s}({n})" for s,n in sorted(v['subs'].items(), key=lambda x:-x[1])[:4])
        bg = '1E2235' if row%2==0 else '161925'
        for c2,(val,color,bold) in enumerate([(cat,'E2E8F0',False),(v['total'],'5B8CFF',True),(top,'94A3B8',False)],1):
            c = ws5.cell(row=row, column=c2, value=val)
            c.font = Font(name='Arial', bold=bold, color=color, size=10)
            c.fill = fill(bg); c.alignment = ca() if c2<=2 else ca('left',wrap=True); c.border = border()
        ws5.row_dimensions[row].height = 16; row += 1
    bar3 = BarChart(); bar3.type='bar'; bar3.title='問題大類件數'; bar3.style=10; bar3.width=16; bar3.height=12
    cats_ref = Reference(ws5, min_col=1, min_row=3, max_row=2+len(sorted_cats))
    data_ref = Reference(ws5, min_col=2, min_row=2, max_row=2+len(sorted_cats))
    bar3.add_data(data_ref, titles_from_data=True); bar3.set_categories(cats_ref)
    ws5.add_chart(bar3, 'E2')

    # ===== ⑥ 處理狀態 =====
    ws6 = wb.create_sheet('⑥ 處理狀態總覽')
    ws6.sheet_view.showGridLines = False
    title_row(ws6, 1, f'📋 處理狀態總覽　｜　{label}', 4)
    set_hdr(ws6, 2, ['處理狀態','件數','處理人員','重點說明'])
    ws6.column_dimensions['A'].width = 14; ws6.column_dimensions['B'].width = 8
    ws6.column_dimensions['C'].width = 22; ws6.column_dimensions['D'].width = 78
    status_groups = {}
    for r in records:
        k = r.get('status') or '未知'
        status_groups.setdefault(k, []).append(r)
    row = 3
    for st, rows in sorted(status_groups.items(), key=lambda x:-len(x[1])):
        handlers = '、'.join(set(r['handler'] for r in rows if r.get('handler')))
        notes = build_status_focus(rows)
        bg = '1E2235' if row%2==0 else '161925'
        for c2, val in enumerate([st,len(rows),handlers,notes],1):
            c = ws6.cell(row=row, column=c2, value=val)
            c.font = Font(name='Arial', bold=(c2==1), color=STATUS_COLORS.get(st,'E2E8F0') if c2==1 else 'E2E8F0', size=10)
            c.fill = fill(bg); c.alignment = ca() if c2<=2 else ca('left',wrap=True); c.border = border()
        ws6.row_dimensions[row].height = max(45, len(notes.split('\n'))*22); row += 1
    chart_row = row+1
    ws6.cell(row=chart_row, column=1, value='狀態'); ws6.cell(row=chart_row, column=2, value='件數')
    for i,(st,rows) in enumerate(sorted(status_groups.items(), key=lambda x:-len(x[1])),1):
        ws6.cell(row=chart_row+i, column=1, value=st); ws6.cell(row=chart_row+i, column=2, value=len(rows))
    pie3 = PieChart(); pie3.title='處理狀態分佈'; pie3.style=10; pie3.width=14; pie3.height=10
    lb3 = Reference(ws6, min_col=1, min_row=chart_row+1, max_row=chart_row+len(status_groups))
    d3 = Reference(ws6, min_col=2, min_row=chart_row, max_row=chart_row+len(status_groups))
    pie3.add_data(d3, titles_from_data=True); pie3.set_categories(lb3)
    ws6.add_chart(pie3, 'G2')

    return wb

# ===== LINE OA webhook helpers =====
LINE_WEBHOOK_SOURCE = 'line_webhook'
LINE_MESSAGES_SCHEMA = (os.environ.get('LINE_MESSAGES_SCHEMA') or 'staging').strip()
LINE_MESSAGES_TABLE = (os.environ.get('LINE_MESSAGES_TABLE') or 'line_messages').strip()
LINE_API_BASE_URL = 'https://api.line.me/v2/bot'

def get_required_env(name):
    value = (os.environ.get(name) or '').strip()
    if not value:
        raise RuntimeError(f'Missing required environment variable: {name}')
    return value

def normalize_line_message(text):
    return ' '.join((text or '').strip().split())

def verify_line_signature(raw_body, signature):
    if not signature:
        return False
    channel_secret = get_required_env('LINE_CHANNEL_SECRET')
    digest = hmac.new(channel_secret.encode('utf-8'), raw_body, hashlib.sha256).digest()
    expected = base64.b64encode(digest).decode('utf-8')
    return hmac.compare_digest(expected, signature)

def line_timestamp_to_iso(timestamp_ms):
    if timestamp_ms is None:
        return None
    try:
        timestamp_seconds = int(timestamp_ms) / 1000
        return datetime.fromtimestamp(timestamp_seconds, tz=timezone.utc).isoformat()
    except Exception:
        return None

def line_api_get(path):
    access_token = (os.environ.get('LINE_CHANNEL_ACCESS_TOKEN') or '').strip()
    if not access_token:
        return None
    response = requests.get(
        f'{LINE_API_BASE_URL}{path}',
        headers={'Authorization': f'Bearer {access_token}'},
        timeout=10
    )
    if response.status_code >= 400:
        app.logger.warning('LINE API GET %s failed: %s %s', path, response.status_code, response.text[:300])
        return None
    return response.json()

def resolve_line_sender_name(source):
    source_type = source.get('type') or 'line'
    user_id = source.get('userId')
    group_id = source.get('groupId')
    room_id = source.get('roomId')

    if source_type == 'group' and group_id:
        summary = line_api_get(f'/group/{group_id}/summary')
        if summary and summary.get('groupName'):
            return summary['groupName']
        return 'LINE 群組'

    if source_type == 'user' and user_id:
        profile = line_api_get(f'/profile/{user_id}')
        if profile and profile.get('displayName'):
            return profile['displayName']
        return 'LINE 用戶'

    if source_type == 'room' and room_id:
        return 'LINE 多人聊天室'

    return source_type or 'line'

def build_line_duplicate_hash(sender_id, sender_name, received_at, normalized_message, message_id=''):
    key_parts = [
        LINE_WEBHOOK_SOURCE,
        sender_id or sender_name or '',
        received_at or '',
        normalized_message or '',
        message_id or ''
    ]
    key = '|'.join(key_parts)
    return hashlib.sha256(key.encode('utf-8')).hexdigest() if normalized_message or message_id else None

def get_supabase_headers(schema=None):
    service_key = get_required_env('SUPABASE_SERVICE_ROLE_KEY')
    headers = {
        'apikey': service_key,
        'Authorization': f'Bearer {service_key}',
        'Content-Type': 'application/json'
    }
    if schema:
        headers['Accept-Profile'] = schema
        headers['Content-Profile'] = schema
    return headers

def supabase_rest_url(table, query=''):
    base_url = get_required_env('SUPABASE_URL').rstrip('/')
    if base_url.endswith('/rest/v1'):
        suffix = f'?{query}' if query else ''
        return f'{base_url}/{table}{suffix}'
    suffix = f'?{query}' if query else ''
    return f'{base_url}/rest/v1/{table}{suffix}'

def supabase_request(method, table, *, schema=None, query='', payload=None):
    response = requests.request(
        method,
        supabase_rest_url(table, query),
        headers=get_supabase_headers(schema),
        data=json.dumps(payload) if payload is not None else None,
        timeout=15
    )
    if response.status_code >= 400:
        raise RuntimeError(f'Supabase {method} {table} failed: {response.status_code} {response.text[:500]}')
    if not response.text:
        return None
    return response.json()

def line_message_exists(duplicate_hash):
    if not duplicate_hash:
        return False
    query = f'duplicate_hash=eq.{duplicate_hash}&select=id&limit=1'
    rows = supabase_request('GET', LINE_MESSAGES_TABLE, schema=LINE_MESSAGES_SCHEMA, query=query)
    return bool(rows)

def insert_line_message(row):
    return supabase_request(
        'POST',
        LINE_MESSAGES_TABLE,
        schema=LINE_MESSAGES_SCHEMA,
        query='select=id',
        payload=row
    )

def log_line_activity(action, detail):
    try:
        payload = {
            'case_id': None,
            'action': action,
            'changed_by': 'LINE Webhook',
            'changed_at': datetime.now(timezone.utc).isoformat(),
            'detail': detail
        }
        supabase_request('POST', 'activity_log', schema='public', payload=payload)
    except Exception as exc:
        app.logger.warning('LINE activity_log write failed: %s', exc)

def build_line_message_row(event):
    source = event.get('source') or {}
    message = event.get('message') or {}
    message_type = message.get('type')
    sender_id = source.get('userId') or source.get('groupId') or source.get('roomId')
    sender_name = resolve_line_sender_name(source)
    received_at = line_timestamp_to_iso(event.get('timestamp'))

    if message_type == 'text':
        raw_message = message.get('text') or ''
        status = 'pending' if raw_message.strip() else 'error'
    else:
        raw_message = json.dumps({
            'event_type': event.get('type'),
            'message_type': message_type,
            'message_id': message.get('id')
        }, ensure_ascii=False, separators=(',', ':'))
        status = 'error'

    normalized_message = normalize_line_message(raw_message)
    duplicate_hash = build_line_duplicate_hash(
        sender_id,
        sender_name,
        received_at,
        normalized_message,
        message.get('id') or event.get('webhookEventId') or ''
    )

    return {
        'source': LINE_WEBHOOK_SOURCE,
        'sender_name': sender_name,
        'sender_id': sender_id,
        'raw_message': raw_message,
        'normalized_message': normalized_message or None,
        'received_at': received_at,
        'status': status,
        'operator_id': 'line_webhook',
        'duplicate_hash': duplicate_hash
    }

def summarize_line_event_for_log(event):
    message = event.get('message') or {}
    source = event.get('source') or {}
    parts = [
        f'event_type={event.get("type") or "-"}',
        f'message_type={message.get("type") or "-"}',
        f'source_type={source.get("type") or "-"}',
        f'webhook_event_id={event.get("webhookEventId") or "-"}'
    ]
    return '; '.join(parts)

def process_line_webhook_events(events):
    summary = {'received': len(events), 'inserted': 0, 'duplicates': 0, 'skipped': 0, 'errors': 0}
    for event in events:
        if event.get('type') != 'message':
            summary['skipped'] += 1
            log_line_activity('line_webhook_event_skipped', summarize_line_event_for_log(event))
            continue
        try:
            row = build_line_message_row(event)
            duplicate_hash = row.get('duplicate_hash')
            if duplicate_hash and line_message_exists(duplicate_hash):
                summary['duplicates'] += 1
                log_line_activity('line_webhook_duplicate_skipped', f'duplicate_hash={duplicate_hash}')
                continue
            insert_line_message(row)
            summary['inserted'] += 1
            log_line_activity('line_webhook_message_imported', f'status={row["status"]}; source={LINE_WEBHOOK_SOURCE}')
        except Exception as exc:
            summary['errors'] += 1
            app.logger.exception('LINE webhook event processing failed')
            log_line_activity('line_webhook_event_failed', str(exc)[:300])
            raise
    return summary

# ===== API ROUTES =====
@app.route('/health', methods=['GET'])
def health():
    return jsonify({'status': 'ok', 'time': datetime.now().isoformat()})

@app.route('/api/line/webhook', methods=['POST'])
def line_webhook():
    raw_body = request.get_data()
    signature = request.headers.get('x-line-signature', '')
    try:
        if not verify_line_signature(raw_body, signature):
            return jsonify({'error': 'invalid LINE signature'}), 403
    except RuntimeError as exc:
        return jsonify({'error': str(exc)}), 500

    try:
        payload = json.loads(raw_body.decode('utf-8') or '{}')
    except Exception:
        return jsonify({'error': 'invalid JSON payload'}), 400

    events = payload.get('events', [])
    if not isinstance(events, list):
        return jsonify({'error': 'invalid LINE events payload'}), 400

    try:
        summary = process_line_webhook_events(events)
        return jsonify({'status': 'ok', **summary})
    except RuntimeError as exc:
        return jsonify({'error': str(exc)}), 500
    except Exception as exc:
        app.logger.exception('LINE webhook request failed')
        return jsonify({'error': str(exc)}), 500

@app.route('/api/line/config-check', methods=['GET'])
def line_config_check():
    try:
        supabase_url = get_required_env('SUPABASE_URL')
        parsed = urlparse(supabase_url)
        return jsonify({
            'status': 'ok',
            'supabase_scheme': parsed.scheme,
            'supabase_host': parsed.netloc,
            'line_messages_schema': LINE_MESSAGES_SCHEMA,
            'line_messages_table': LINE_MESSAGES_TABLE,
            'has_line_channel_secret': bool((os.environ.get('LINE_CHANNEL_SECRET') or '').strip()),
            'has_supabase_service_role_key': bool((os.environ.get('SUPABASE_SERVICE_ROLE_KEY') or '').strip())
        })
    except RuntimeError as exc:
        return jsonify({'status': 'error', 'error': str(exc)}), 500

@app.route('/api/line/supabase-check', methods=['GET'])
def line_supabase_check():
    try:
        rows = supabase_request(
            'GET',
            LINE_MESSAGES_TABLE,
            schema=LINE_MESSAGES_SCHEMA,
            query='select=id&limit=1'
        )
        return jsonify({
            'status': 'ok',
            'line_messages_schema': LINE_MESSAGES_SCHEMA,
            'line_messages_table': LINE_MESSAGES_TABLE,
            'row_count_sample': len(rows or [])
        })
    except Exception as exc:
        supabase_url = (os.environ.get('SUPABASE_URL') or '').strip()
        parsed = urlparse(supabase_url)
        return jsonify({
            'status': 'error',
            'supabase_host': parsed.netloc,
            'line_messages_schema': LINE_MESSAGES_SCHEMA,
            'line_messages_table': LINE_MESSAGES_TABLE,
            'error_type': type(exc).__name__,
            'error': str(exc)[:500]
        }), 500

@app.route('/weekly-report', methods=['POST'])
def weekly_report():
    try:
        data = request.json
        records = data.get('records', [])
        from_date = data.get('from', '')
        to_date = data.get('to', '')
        all_records = data.get('all_records', records)  # 全部資料，預設用 records
        if not records:
            return jsonify({'error': '無資料'}), 400
        wb = generate_weekly(records, from_date, to_date, all_records)
        buf = io.BytesIO()
        wb.save(buf); buf.seek(0)
        filename = f'週報_{from_date}_{to_date}.xlsx'
        return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                        as_attachment=True, download_name=filename)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/monthly-report', methods=['POST'])
def monthly_report():
    try:
        data = request.json
        records = data.get('records', [])
        from_date = data.get('from', '')
        to_date = data.get('to', '')
        if not records:
            return jsonify({'error': '無資料'}), 400
        wb = generate_monthly(records, from_date, to_date)
        buf = io.BytesIO()
        wb.save(buf); buf.seek(0)
        filename = f'月報_{from_date}_{to_date}.xlsx'
        return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                        as_attachment=True, download_name=filename)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port)
